# cf-dev/cf_src/appsinn/cf_operations/tests.py

"""Tests for operations domain models (events, sermons, zonal attendance)."""

from django.core.exceptions import ValidationError
from django.test import TestCase
from django.utils import timezone

from cf_users.models import Branch, Organization, User

from .models import (
    AttendanceRecord,
    AttendanceSeat,
    Event,
    EventSession,
    Sermon,
)


class OperationsFixturesMixin:
    def setUp(self):
        self.org = Organization.objects.create(
            name="Ops Org",
            address="Street",
            city="Accra",
            country="GH",
        )
        self.branch = Branch.objects.create(
            organization=self.org,
            name="Main",
            address="Street",
            city="Accra",
            country="GH",
            is_default=True,
        )
        self.org_b = Organization.objects.create(
            name="Other Church",
            address="Road",
            city="Kumasi",
            country="GH",
        )
        self.branch_b = Branch.objects.create(
            organization=self.org_b,
            name="Other Main",
            address="Road",
            city="Kumasi",
            country="GH",
            is_default=True,
        )
        self.user = User.objects.create_user(
            username="viewer1",
            email="viewer1@example.com",
            password="ComplexPass123!",
            first_name="View",
            last_name="One",
            phone_number="+233201000001",
            address="Addr",
            city="Accra",
            country="GH",
        )
        self.user_b = User.objects.create_user(
            username="viewer2",
            email="viewer2@example.com",
            password="ComplexPass123!",
            first_name="View",
            last_name="Two",
            phone_number="+233201000002",
            address="Addr",
            city="Kumasi",
            country="GH",
        )


class EventModelTests(OperationsFixturesMixin, TestCase):
    def test_create_event_defaults_public(self):
        start = timezone.now()
        event = Event.objects.create(
            branch=self.branch,
            title="Sunday Service",
            start_time=start,
            end_time=start + timezone.timedelta(hours=2),
        )
        self.assertEqual(str(event), "Sunday Service")
        self.assertEqual(event.visibility, Event.VISIBILITY_PUBLIC)
        self.assertEqual(event.event_type, "SERVICE")

    def test_branch_only_visibility_choice(self):
        self.assertIn(
            Event.VISIBILITY_BRANCH,
            dict(Event.VISIBILITY_CHOICES),
        )
        event = Event.objects.create(
            branch=self.branch,
            title="Branch meeting",
            visibility=Event.VISIBILITY_BRANCH,
        )
        self.assertEqual(event.visibility, "BRANCH")

    def test_outreach_event_type(self):
        start = timezone.now()
        event = Event.objects.create(
            branch=self.branch,
            title="Street Outreach",
            event_type="OUTREACH",
            start_time=start,
            end_time=start + timezone.timedelta(hours=3),
            visibility=Event.VISIBILITY_PUBLIC,
        )
        self.assertEqual(event.event_type, "OUTREACH")


class SermonGuestSpeakerTests(OperationsFixturesMixin, TestCase):
    def test_guest_speaker_without_member(self):
        sermon = Sermon(
            branch=self.branch,
            title="Grace Abounds",
            guest_speaker_name="John Doe",
            guest_speaker_title="Pastor",
            guest_speaker_church="Visiting Assembly",
        )
        sermon.full_clean()
        sermon.save()
        self.assertIsNone(sermon.speaker_id)
        self.assertIn("John Doe", sermon.get_speaker_display())
        self.assertIn("Visiting Assembly", sermon.get_speaker_display())

    def test_requires_member_or_guest_name(self):
        sermon = Sermon(branch=self.branch, title="Untitled")
        with self.assertRaises(ValidationError) as ctx:
            sermon.full_clean()
        self.assertIn("guest_speaker_name", ctx.exception.message_dict)


class VisibilityQueryTests(OperationsFixturesMixin, TestCase):
    def setUp(self):
        super().setUp()
        start = timezone.now()
        end = start + timezone.timedelta(hours=1)
        self.public_event = Event.objects.create(
            branch=self.branch,
            title="Public Revival",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_PUBLIC,
        )
        self.private_event = Event.objects.create(
            branch=self.branch,
            title="Leaders Only",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_ORGANIZATION,
        )
        self.other_public = Event.objects.create(
            branch=self.branch_b,
            title="Other Public",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_PUBLIC,
        )
        self.other_private = Event.objects.create(
            branch=self.branch_b,
            title="Other Private",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_ORGANIZATION,
        )

    def test_anonymous_sees_only_public(self):
        qs = Event.objects.visible_to(None)
        titles = set(qs.values_list("title", flat=True))
        self.assertIn("Public Revival", titles)
        self.assertIn("Other Public", titles)
        self.assertNotIn("Leaders Only", titles)
        self.assertNotIn("Other Private", titles)

    def test_authenticated_user_sees_public_platform_wide(self):
        # user has no org membership — still sees all PUBLIC content
        qs = Event.objects.visible_to(self.user)
        titles = set(qs.values_list("title", flat=True))
        self.assertIn("Public Revival", titles)
        self.assertIn("Other Public", titles)
        self.assertNotIn("Leaders Only", titles)
        self.assertNotIn("Other Private", titles)


class AttendanceExcelModelTests(OperationsFixturesMixin, TestCase):
    def test_record_seat_excel_fields_no_prepopulate(self):
        event = Event.objects.create(
            branch=self.branch,
            title="Zonal service June",
            event_type="SERVICE",
        )
        record = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=1,
            month=6,
            code="1",
            centre_name="His Presence",
            leader="Mrs. Star Innocent Adukwu",
            address="Iron City 18 Soursop Street",
            location_provider="Cell host",
        )
        # No auto-created weeks 1–5; seat is optional and created on demand.
        self.assertFalse(hasattr(record, "week_stats"))
        with self.assertRaises(AttendanceSeat.DoesNotExist):
            _ = record.seat

        seat = AttendanceSeat.objects.create(
            record=record,
            male_adults=3,
            female_adults=2,
            male_children=2,
            female_children=1,
            testimonies=4,
        )
        seat.refresh_from_db()
        self.assertEqual(seat.total, 8)  # 3+2+2+1
        self.assertEqual(record.week, 1)

    def test_sheet_context_matches_excel_layout(self):
        from .attendance_report import build_sheet_from_records

        event = Event.objects.create(
            branch=self.branch, title="Sheet event", event_type="SERVICE"
        )
        r1 = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=1,
            month=6,
            code="1",
            centre_name="His Presence",
            leader="Leader A",
            address="Loc A",
        )
        AttendanceSeat.objects.create(
            record=r1,
            male_adults=3,
            female_adults=2,
            male_children=2,
            female_children=1,
            testimonies=4,
        )

        qs = AttendanceRecord.objects.filter(pk=r1.pk).select_related("seat")
        sheet = build_sheet_from_records(
            qs,
            assembly_name="KASOA ASSEMBLY",
            zone_name="ZONE 13",
            report_title="ZONAL REPORT JUNE 2026",
            coordinator_name="Coord",
            week_labels=["WEEK 1 - 01/08/2026", "WEEK 2", "WEEK 3", "WEEK 4", "WEEK 5"],
        )
        self.assertEqual(sheet["assembly_name"], "KASOA ASSEMBLY")
        self.assertEqual(sheet["zone_name"], "ZONE 13")
        self.assertEqual(len(sheet["week_labels"]), 1)
        self.assertEqual(sheet["week_numbers"], [1])
        self.assertEqual(sheet["rows"][0]["centre_name"], "His Presence")
        self.assertEqual(sheet["rows"][0]["weeks"][0]["male_adults"], 3)
        self.assertEqual(sheet["rows"][0]["weeks"][0]["total"], 8)
        self.assertEqual(sheet["week_totals"][0]["total"], 8)
        self.assertEqual(sheet["total_cells"], 1)
        self.assertEqual(sheet["active_cells"], 1)
        self.assertEqual(sheet["grand"]["total"], 8)
        self.assertEqual(sheet["month_label"], "June")

    def test_preview_shows_zone_from_record(self):
        from cf_people.models import Zone

        from .attendance_report import build_sheet_from_records

        event = Event.objects.create(
            branch=self.branch, title="Zone event", event_type="SERVICE"
        )
        zone = Zone.objects.create(branch=self.branch, name="ZONE 13", code="Z13")
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            zone=zone,
            week=2,
            code="1",
            centre_name="Cell A",
        )
        AttendanceSeat.objects.create(
            record=rec, male_adults=1, female_adults=1
        )
        qs = AttendanceRecord.objects.filter(pk=rec.pk).select_related("zone", "seat")
        sheet = build_sheet_from_records(qs)
        self.assertEqual(sheet["zone_name"], "ZONE 13")
        self.assertEqual(sheet["rows"][0]["zone_name"], "ZONE 13")
        self.assertEqual(sheet["week_numbers"], [2])
        self.assertEqual(len(sheet["rows"][0]["weeks"]), 1)
        self.assertEqual(sheet["rows"][0]["weeks"][0]["total"], 2)
        self.assertTrue(sheet["week_labels"][0].startswith("WEEK 2"))

    def test_attendance_record_event_zone_subgroup(self):
        from cf_people.models import SubBranch, Zone

        event = Event.objects.create(
            branch=self.branch, title="Sunday service", event_type="SERVICE"
        )
        zone = Zone.objects.create(branch=self.branch, name="ZONE 13", code="Z13")
        cell = SubBranch.objects.create(
            zone=zone, branch=self.branch, name="His Presence", group_type="CELL"
        )
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            zone=zone,
            subgroup=cell,
            code="1",
        )
        self.assertEqual(rec.event_id, event.pk)
        self.assertEqual(rec.zone_id, zone.pk)
        self.assertEqual(rec.subgroup_id, cell.pk)
        # Without fill_from_scope, override stays blank; display falls back to sub group.
        self.assertEqual(rec.centre_name, "")
        self.assertEqual(rec.get_display_centre_name(), "His Presence")
        self.assertIsNone(rec.week)

    def test_attendance_month_datetime_leader_text(self):
        from datetime import datetime

        from django.utils import timezone as dj_tz

        event = Event.objects.create(
            branch=self.branch, title="Month check", event_type="SERVICE"
        )
        when = dj_tz.make_aware(datetime(2026, 6, 15, 9, 30))
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=2,
            month=6,
            attendance_at=when,
            code="C-01",
            leader="Deacon Host",
            centre_name="Cell One",
            address="Hall A",
            phone_number="+233201111111",
            location_provider="Host Family",
        )
        rec.full_clean()
        self.assertEqual(rec.get_month_display(), "June")
        self.assertEqual(rec.leader, "Deacon Host")
        self.assertEqual(rec.location_provider, "Host Family")
        self.assertEqual(rec.attendance_at, when)
        self.assertEqual(rec.address, "Hall A")
        self.assertEqual(rec.phone_number, "+233201111111")

        from .attendance_report import build_sheet_from_records

        AttendanceSeat.objects.create(record=rec, male_adults=2, female_adults=3)
        sheet = build_sheet_from_records(
            AttendanceRecord.objects.filter(pk=rec.pk).select_related("seat", "zone")
        )
        self.assertEqual(sheet["rows"][0]["leader_name"], "Deacon Host")
        self.assertEqual(sheet["rows"][0]["address"], "Hall A")
        self.assertEqual(sheet["rows"][0]["contact"], "+233201111111")
        self.assertEqual(sheet["rows"][0]["code"], "C-01")
        self.assertEqual(sheet["week_numbers"], [2])
        self.assertEqual(sheet["rows"][0]["weeks"][0]["total"], 5)
        self.assertEqual(sheet["month_label"], "June")

    def test_sheet_header_scope_and_leader_not_overwritten(self):
        from types import SimpleNamespace

        from .attendance_report import (
            build_sheet_from_records,
            format_report_date,
            resolve_sheet_header,
        )

        org = SimpleNamespace(name="City Church", trade_name="", leader="Pastor Org")
        branch = SimpleNamespace(
            name="Kasoa Branch", organization=org, leader="Elder Branch"
        )
        zone = SimpleNamespace(name="ZONE 13", branch=branch, leader="Coord Zone")
        subgroup = SimpleNamespace(
            name="His Presence", zone=zone, branch=branch, leader="Host Cell"
        )

        org_header = resolve_sheet_header(organization=org, report_date="01 June 2026")
        self.assertEqual(org_header["organization_name"], "City Church")
        self.assertEqual(org_header["branch_name"], "")
        self.assertFalse(org_header["show_zone"])
        self.assertFalse(org_header["show_subgroup"])
        self.assertEqual(org_header["leader_name"], "Pastor Org")
        self.assertEqual(org_header["report_date"], "01 June 2026")

        branch_header = resolve_sheet_header(branch=branch)
        self.assertEqual(branch_header["organization_name"], "City Church")
        self.assertEqual(branch_header["branch_name"], "Kasoa Branch")
        self.assertFalse(branch_header["show_zone"])
        self.assertEqual(branch_header["leader_name"], "Elder Branch")

        zone_header = resolve_sheet_header(zone=zone)
        self.assertTrue(zone_header["show_zone"])
        self.assertEqual(zone_header["zone_name"], "ZONE 13")
        self.assertFalse(zone_header["show_subgroup"])
        self.assertEqual(zone_header["leader_name"], "Coord Zone")

        cell_header = resolve_sheet_header(subgroup=subgroup)
        self.assertTrue(cell_header["show_zone"])
        self.assertTrue(cell_header["show_subgroup"])
        self.assertEqual(cell_header["subgroup_name"], "His Presence")
        self.assertEqual(cell_header["leader_name"], "Host Cell")

        self.assertEqual(
            format_report_date({"date": "2026-06-15"}),
            "15 June 2026",
        )
        self.assertEqual(
            format_report_date({"month": "6", "year": "2026"}),
            "June 2026",
        )

        event = Event.objects.create(
            branch=self.branch, title="Header event", event_type="SERVICE"
        )
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=1,
            leader="Row Leader",
            centre_name="Cell A",
        )
        AttendanceSeat.objects.create(record=rec, male_adults=1, female_adults=1)
        sheet = build_sheet_from_records(
            AttendanceRecord.objects.filter(pk=rec.pk).select_related("seat"),
            organization_name="City Church",
            branch_name="Kasoa Branch",
            zone_name="ZONE 13",
            subgroup_name="His Presence",
            report_date="15 June 2026",
            show_zone=True,
            show_subgroup=True,
            header_leader_name="Coord Zone",
        )
        self.assertEqual(sheet["organization_name"], "City Church")
        self.assertEqual(sheet["branch_name"], "Kasoa Branch")
        self.assertEqual(sheet["zone_name"], "ZONE 13")
        self.assertEqual(sheet["subgroup_name"], "His Presence")
        self.assertEqual(sheet["report_date"], "15 June 2026")
        self.assertEqual(sheet["leader_name"], "Coord Zone")
        self.assertEqual(sheet["rows"][0]["leader_name"], "Row Leader")

    def test_week_labels_include_attendance_date(self):
        from datetime import datetime

        from django.utils import timezone as dj_tz

        from .attendance_report import build_sheet_from_records

        event = Event.objects.create(
            branch=self.branch, title="Dated weeks", event_type="SERVICE"
        )
        w1 = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=1,
            attendance_at=dj_tz.make_aware(datetime(2026, 8, 1, 9, 0)),
            centre_name="Cell A",
        )
        w2 = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=2,
            attendance_at=dj_tz.make_aware(datetime(2026, 8, 8, 9, 0)),
            centre_name="Cell A",
        )
        AttendanceSeat.objects.create(record=w1, male_adults=1)
        AttendanceSeat.objects.create(record=w2, male_adults=1)
        sheet = build_sheet_from_records(
            AttendanceRecord.objects.filter(event=event).select_related("seat")
        )
        self.assertEqual(sheet["week_labels"][0], "WEEK 1 - 2026-08-01")
        self.assertEqual(sheet["week_labels"][1], "WEEK 2 - 2026-08-08")
        self.assertEqual(len(sheet["week_labels"]), 2)
        self.assertEqual(sheet["week_numbers"], [1, 2])

    def test_progressive_weeks_omit_empty_rows_and_empty_weeks(self):
        from .attendance_export import export_sheet_xlsx
        from .attendance_report import build_sheet_from_records

        event = Event.objects.create(
            branch=self.branch, title="Week three only", event_type="SERVICE"
        )
        empty = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=1,
            centre_name="Empty cell",
        )
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            week=3,
            centre_name="Active cell",
        )
        AttendanceSeat.objects.create(record=rec, male_adults=4, female_adults=1)
        sheet = build_sheet_from_records(
            AttendanceRecord.objects.filter(event=event).select_related("seat")
        )
        self.assertEqual(len(sheet["rows"]), 1)
        self.assertEqual(sheet["rows"][0]["centre_name"], "Active cell")
        self.assertEqual(sheet["week_numbers"], [3])
        self.assertEqual(len(sheet["week_labels"]), 1)
        self.assertTrue(sheet["week_labels"][0].startswith("WEEK 3"))
        self.assertEqual(sheet["rows"][0]["weeks"][0]["male_adults"], 4)
        self.assertEqual(empty.centre_name, "Empty cell")

        payload = export_sheet_xlsx(sheet)
        from io import BytesIO

        from openpyxl import load_workbook

        wb = load_workbook(BytesIO(payload))
        ws = wb.active
        # Week header + col headers + one data row should be bordered.
        data_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            if row[0].value == "CODE":
                self.assertIsNotNone(row[0].border.left.style)
            if row[1].value == "Active cell":
                data_row = row
        self.assertIsNotNone(data_row)
        self.assertIsNotNone(data_row[0].border.left.style)
        # MA (col 6) same width as FA (col 7)
        self.assertEqual(
            ws.column_dimensions["F"].width,
            ws.column_dimensions["G"].width,
        )

    def test_fill_from_scope_hierarchy(self):
        from cf_people.models import Member, SubBranch, Zone

        event = Event.objects.create(
            branch=self.branch, title="Scope fill", event_type="SERVICE"
        )
        # Branch-level leader + address
        branch_leader = Member.objects.create(
            branch=self.branch,
            organization=self.org,
            user=self.user,
            membership_status="ACTIVE",
        )
        self.branch.leader = branch_leader
        self.branch.address = "Branch Street 1"
        self.branch.save()

        # Branch only → branch details
        self.branch.code = "BR01"
        self.branch.save(update_fields=["code", "leader", "address"])

        rec_branch = AttendanceRecord(
            event=event,
            branch=self.branch,
            fill_from_scope=True,
        )
        rec_branch.save()
        rec_branch.refresh_from_db()
        self.assertEqual(rec_branch.get_scope_level(), "branch")
        self.assertEqual(rec_branch.code, "BR01")
        self.assertEqual(rec_branch.centre_name, self.branch.name)
        self.assertEqual(rec_branch.leader, str(branch_leader))
        self.assertEqual(rec_branch.address, "Branch Street 1")
        self.assertEqual(
            rec_branch.phone_number, str(self.user.phone_number)
        )

        # Zone overrides branch
        zone = Zone.objects.create(
            branch=self.branch,
            name="ZONE 7",
            code="Z7",
            address="Zone Road 7",
            leader=branch_leader,
        )
        rec_zone = AttendanceRecord(
            event=event,
            branch=self.branch,
            zone=zone,
            fill_from_scope=True,
        )
        rec_zone.save()
        rec_zone.refresh_from_db()
        self.assertEqual(rec_zone.get_scope_level(), "zone")
        self.assertEqual(rec_zone.code, "Z7")
        self.assertEqual(rec_zone.centre_name, "ZONE 7")
        self.assertEqual(rec_zone.address, "Zone Road 7")

        # Sub group overrides zone
        cell = SubBranch.objects.create(
            zone=zone,
            branch=self.branch,
            name="Grace Cell",
            code="CELL-3",
            group_type="CELL",
            address="Cell Lane 3",
            location_provider="Sister Ama",
            leader=branch_leader,
        )
        rec_cell = AttendanceRecord(
            event=event,
            branch=self.branch,
            zone=zone,
            subgroup=cell,
            fill_from_scope=True,
        )
        rec_cell.save()
        rec_cell.refresh_from_db()
        self.assertEqual(rec_cell.get_scope_level(), "subgroup")
        self.assertEqual(rec_cell.code, "CELL-3")
        self.assertEqual(rec_cell.centre_name, "Grace Cell")
        self.assertEqual(rec_cell.address, "Cell Lane 3")
        self.assertEqual(rec_cell.location_provider, "Sister Ama")

        # Manual override wins over scope defaults when fill_from_scope is off
        rec_override = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            zone=zone,
            subgroup=cell,
            code="CUSTOM",
            centre_name="Custom Centre",
            leader="Custom Leader",
            fill_from_scope=False,
        )
        self.assertEqual(rec_override.get_display_code(), "CUSTOM")
        self.assertEqual(rec_override.get_display_centre_name(), "Custom Centre")
        self.assertEqual(rec_override.get_display_leader(), "Custom Leader")
        # Empty address falls back to sub group
        self.assertEqual(rec_override.get_display_address(), "Cell Lane 3")
        # Empty code on another record falls back to sub group code
        rec_blank_code = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            zone=zone,
            subgroup=cell,
            fill_from_scope=False,
        )
        self.assertEqual(rec_blank_code.code, "")
        self.assertEqual(rec_blank_code.get_display_code(), "CELL-3")

    def test_event_session_and_attendance_link(self):
        event = Event.objects.create(
            branch=self.branch,
            title="Sunday Combined",
            event_type="SERVICE",
        )
        session = EventSession.objects.create(
            event=event,
            name="1st Service",
            sort_order=1,
        )
        self.assertEqual(session.branch_id, self.branch.pk)
        self.assertIn("1st Service", str(session))
        rec = AttendanceRecord.objects.create(
            event=event,
            session=session,
            branch=self.branch,
            centre_name="Main hall",
        )
        self.assertEqual(rec.session_id, session.pk)
        self.assertEqual(rec.session.event_id, event.pk)

    def test_event_session_weekday_and_time_fields(self):
        from datetime import datetime, time

        from django.core.exceptions import ValidationError
        from django.utils import timezone as dj_tz

        event = Event.objects.create(
            branch=self.branch,
            title="Sunday Services",
            event_type="SERVICE",
        )
        check_in_start = dj_tz.make_aware(datetime(2026, 8, 9, 6, 30))
        check_in_end = dj_tz.make_aware(datetime(2026, 8, 9, 8, 45))
        session = EventSession.objects.create(
            event=event,
            name="1st Service",
            start_day=EventSession.WEEKDAY_SUNDAY,
            start_time=time(7, 0),
            end_day=EventSession.WEEKDAY_SUNDAY,
            end_time=time(9, 0),
            check_in_start=check_in_start,
            check_in_end=check_in_end,
        )
        session.refresh_from_db()

        self.assertEqual(session.start_day, "SUN")
        self.assertEqual(session.start_time, time(7, 0))
        self.assertEqual(session.end_day, "SUN")
        self.assertEqual(session.end_time, time(9, 0))
        self.assertEqual(session.get_start_day_display(), "Sunday")
        self.assertTrue(
            session.is_check_in_open(
                at=dj_tz.make_aware(datetime(2026, 8, 9, 7, 15))
            )
        )
        self.assertFalse(
            session.is_check_in_open(
                at=dj_tz.make_aware(datetime(2026, 8, 9, 9, 30))
            )
        )

        bad = EventSession(
            event=event,
            branch=self.branch,
            name="Broken",
            start_day=EventSession.WEEKDAY_SUNDAY,
            start_time=time(10, 0),
            end_day=EventSession.WEEKDAY_SUNDAY,
            end_time=time(9, 0),
        )
        with self.assertRaises(ValidationError):
            bad.full_clean()


class AttendanceExcelImportTests(OperationsFixturesMixin, TestCase):
    def test_import_from_minimal_workbook(self):
        from io import BytesIO

        from openpyxl import Workbook

        from .attendance_import import import_attendance_from_excel

        wb = Workbook()
        ws = wb.active
        ws["A2"] = "KASOA ASSEMBLY"
        ws["A3"] = "ZONAL REPORT JUNE 2026"
        ws["A4"] = "ZONE 13"
        ws["B5"] = "ZONAL COORDINATOR NAME: Test Coord"
        ws["F5"] = "WEEK 1 - 01/08/2026"
        ws["N5"] = "WEEK 2"
        ws["V5"] = "WEEK 3"
        ws["AD5"] = "WEEK 4"
        ws["AL5"] = "WEEK 5"
        # Centre row
        ws["A7"] = 1
        ws["B7"] = "His Presence"
        ws["C7"] = "Leader A"
        ws["D7"] = "Iron City"
        ws["E7"] = "+233543053095"
        ws["F7"] = 3  # MA
        ws["G7"] = 2  # FA
        ws["H7"] = 2  # MC
        ws["I7"] = 1  # FC
        ws["J7"] = 8  # T
        ws["M7"] = 4  # TS
        ws["F19"] = "TOTAL"

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        result = import_attendance_from_excel(branch=self.branch, file_obj=buf)
        self.assertIsNotNone(result.event)
        self.assertEqual(result.assembly_name, "KASOA ASSEMBLY")
        self.assertEqual(result.zone_name, "ZONE 13")
        # Only weeks with counts become records (week 1 only in this sheet).
        self.assertEqual(result.records_created, 1)
        self.assertEqual(result.seats_created, 1)
        record = AttendanceRecord.objects.get(pk=result.record_ids[0])
        self.assertEqual(record.centre_name, "His Presence")
        self.assertEqual(record.event_id, result.event.pk)
        self.assertEqual(record.week, 1)
        seat = record.seat
        self.assertEqual(seat.male_adults, 3)
        self.assertEqual(seat.total, 8)
        self.assertEqual(seat.testimonies, 4)
        # No empty week 2–5 records.
        self.assertEqual(
            AttendanceRecord.objects.filter(event=result.event).count(), 1
        )


class AttendanceGenerateReportViewTests(OperationsFixturesMixin, TestCase):
    def test_scope_fields_are_autocomplete(self):
        from django.urls import reverse

        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save()
        self.client.force_login(self.user)
        url = reverse("admin:cf_operations_attendancerecord_generate_report")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "admin-autocomplete")
        self.assertContains(response, 'name="organization"')
        self.assertContains(response, 'name="branch"')
        self.assertContains(response, 'name="zone"')
        self.assertContains(response, 'name="subgroup"')
        self.assertContains(response, "admin/js/autocomplete.js")

    def test_generated_sheet_uses_branch_leader_and_header_lines(self):
        from django.urls import reverse

        from cf_people.models import Member, Zone

        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save()
        self.org.leader = Member.objects.create(
            branch=self.branch,
            organization=self.org,
            user=self.user,
            membership_status="ACTIVE",
        )
        self.org.save(update_fields=["leader"])
        self.branch.leader = self.org.leader
        self.branch.save(update_fields=["leader"])
        zone = Zone.objects.create(branch=self.branch, name="ZONE 13", code="Z13")
        event = Event.objects.create(
            branch=self.branch, title="Sunday", event_type="SERVICE"
        )
        rec = AttendanceRecord.objects.create(
            event=event,
            branch=self.branch,
            zone=zone,
            week=1,
            leader="Row Host",
            centre_name="Cell A",
        )
        AttendanceSeat.objects.create(record=rec, male_adults=2, female_adults=2)

        self.client.force_login(self.user)
        url = reverse("admin:cf_operations_attendancerecord_generate_report")
        response = self.client.get(url, {"branch": str(self.branch.pk)})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.org.name)
        self.assertContains(response, self.branch.name)
        self.assertNotContains(response, "sheet-header-line sheet-zone")
        self.assertContains(response, "LEADER NAME:")
        self.assertContains(response, str(self.branch.leader))
        self.assertContains(response, "Row Host")
        self.assertContains(response, 'class="label-cell">GRAND TOTAL')
        self.assertNotContains(response, "label-cell sheet-align-left")

        zonal = self.client.get(url, {"zone": str(zone.pk)})
        self.assertContains(zonal, "sheet-header-line sheet-zone")
        self.assertContains(zonal, self.branch.name)

    def test_changelist_filters_use_autocomplete(self):
        from django.urls import reverse

        self.user.is_staff = True
        self.user.is_superuser = True
        self.user.save()
        self.client.force_login(self.user)
        url = reverse("admin:cf_operations_attendancerecord_changelist")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "admin-autocomplete")
        self.assertContains(response, "cf-select2")


class PortalExploreViewTests(OperationsFixturesMixin, TestCase):
    def setUp(self):
        super().setUp()
        start = timezone.now()
        end = start + timezone.timedelta(hours=2)
        self.public_event = Event.objects.create(
            branch=self.branch,
            title="Public Night of Prayer",
            event_type="SERVICE",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_PUBLIC,
        )
        self.outreach = Event.objects.create(
            branch=self.branch_b,
            title="Street Outreach Accra",
            event_type="OUTREACH",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_PUBLIC,
        )
        self.private_event = Event.objects.create(
            branch=self.branch_b,
            title="Private Leaders Meeting",
            event_type="MEETING",
            start_time=start,
            end_time=end,
            visibility=Event.VISIBILITY_ORGANIZATION,
        )
        self.sermon = Sermon.objects.create(
            branch=self.branch,
            title="Grace for Today",
            guest_speaker_name="Guest Pastor",
            guest_speaker_title="Pastor",
            visibility=Sermon.VISIBILITY_PUBLIC,
        )

    def test_event_list_shows_public_hides_private_other_org(self):
        self.client.force_login(self.user)
        response = self.client.get("/explore/events/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("Public Night of Prayer", content)
        self.assertIn("Street Outreach Accra", content)
        self.assertNotIn("Private Leaders Meeting", content)

    def test_outreach_list_filters_type(self):
        self.client.force_login(self.user)
        response = self.client.get("/explore/outreaches/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("Street Outreach Accra", content)
        self.assertNotIn("Public Night of Prayer", content)

    def test_sermon_detail_guest_speaker(self):
        self.client.force_login(self.user)
        response = self.client.get(f"/explore/sermons/{self.sermon.pk}/")
        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("Grace for Today", content)
        self.assertIn("Guest Pastor", content)

    def test_explore_requires_login(self):
        response = self.client.get("/explore/events/")
        self.assertEqual(response.status_code, 302)
        # Portal login is at / (portal_login); next carries the original path.
        self.assertIn("next=", response.url)
        self.assertIn("/explore/events/", response.url)
