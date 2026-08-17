# cf-dev/cf_src/appsinn/cf_operations/attendance_export.py

"""Export attendance sheet dict to Excel (.xlsx) or PDF."""

from __future__ import annotations

from io import BytesIO
from typing import Any


_WEEK_FILLS = (
    "93C5FD",
    "86EFAC",
    "FDBA74",
    "D8B4FE",
    "5EEAD4",
)


def export_sheet_xlsx(sheet: dict[str, Any]) -> bytes:
    """Build an Excel workbook matching the preview grid (borders + fills)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    labels = list(sheet.get("week_labels") or [])
    week_numbers = list(sheet.get("week_numbers") or [])
    while len(week_numbers) < len(labels):
        week_numbers.append(len(week_numbers) + 1)

    meta_cols = 5
    week_width = 8
    total_cols = meta_cols + week_width * len(labels)
    last_col = max(total_cols, 2)

    thin = Border(
        left=Side(style="thin", color="64748B"),
        right=Side(style="thin", color="64748B"),
        top=Side(style="thin", color="64748B"),
        bottom=Side(style="thin", color="64748B"),
    )
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    fill_white = PatternFill("solid", fgColor="FFFFFF")
    fill_header = PatternFill("solid", fgColor="E2E8F0")
    fill_leader = PatternFill("solid", fgColor="FDE68A")
    fill_alt = PatternFill("solid", fgColor="F8FAFC")
    fill_total_col = PatternFill("solid", fgColor="FEF3C7")
    fill_totals = PatternFill("solid", fgColor="BBF7D0")
    fill_summary = PatternFill("solid", fgColor="F1F5F9")

    font_org = Font(name="Arial", bold=True, size=16, color="0F172A")
    font_branch = Font(name="Arial", bold=True, size=12, color="1E293B")
    font_zone = Font(name="Arial", bold=True, size=11, color="1E3A8A")
    font_date = Font(name="Arial", bold=True, size=11, color="334155")
    font_leader = Font(name="Arial", bold=True, size=10, color="78350F")
    font_week = Font(name="Arial", bold=True, size=10, color="0F172A")
    font_col = Font(name="Arial", bold=True, size=9, color="0F172A")
    font_cell = Font(name="Arial", size=9, color="111827")
    font_totals = Font(name="Arial", bold=True, size=9, color="14532D")
    font_grand = Font(name="Arial", bold=True, size=12, color="1E3A8A")
    font_summary = Font(name="Arial", bold=True, size=10, color="0F172A")

    def _write_title(text, font):
        ws.append([text])
        row_idx = ws.max_row
        if last_col > 1:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=last_col)
        cell = ws.cell(row_idx, 1)
        cell.font = font
        cell.alignment = left
        cell.fill = fill_white
        return row_idx

    _write_title(
        sheet.get("organization_name") or sheet.get("assembly_name") or "",
        font_org,
    )
    if sheet.get("branch_name"):
        _write_title(sheet.get("branch_name"), font_branch)
    if sheet.get("show_zone") and sheet.get("zone_name"):
        _write_title(sheet.get("zone_name"), font_zone)
    if sheet.get("show_subgroup") and sheet.get("subgroup_name"):
        _write_title(sheet.get("subgroup_name"), font_branch)
    if sheet.get("report_date"):
        _write_title(sheet.get("report_date"), font_date)

    leader = sheet.get("leader_name") or sheet.get("coordinator_name") or ""
    week_row_vals = [f"LEADER NAME: {leader}"] + [""] * (meta_cols - 1)
    for lab in labels:
        week_row_vals.append(lab)
        week_row_vals.extend([""] * (week_width - 1))
    ws.append(week_row_vals)
    week_row_idx = ws.max_row
    if meta_cols > 1:
        ws.merge_cells(
            start_row=week_row_idx,
            start_column=1,
            end_row=week_row_idx,
            end_column=meta_cols,
        )
    for i, lab in enumerate(labels):
        start = meta_cols + 1 + i * week_width
        end = start + week_width - 1
        if end > start:
            ws.merge_cells(
                start_row=week_row_idx,
                start_column=start,
                end_row=week_row_idx,
                end_column=end,
            )
        wn = week_numbers[i] if i < len(week_numbers) else i + 1
        color = _WEEK_FILLS[(int(wn) - 1) % len(_WEEK_FILLS)]
        fill = PatternFill("solid", fgColor=color)
        for col in range(start, end + 1):
            cell = ws.cell(week_row_idx, col)
            cell.fill = fill
            cell.font = font_week
            cell.alignment = center
            cell.border = thin
    for col in range(1, meta_cols + 1):
        cell = ws.cell(week_row_idx, col)
        cell.fill = fill_leader
        cell.font = font_leader
        cell.alignment = left
        cell.border = thin

    col_headers = ["CODE", "CENTRE NAME", "LEADER'S NAME", "ADDRESS", "CONTACT"]
    for _ in labels:
        col_headers.extend(["MA", "FA", "MC", "FC", "T", "N/C", "F/T", "TS"])
    ws.append(col_headers)
    header_row_idx = ws.max_row
    for col in range(1, last_col + 1):
        cell = ws.cell(header_row_idx, col)
        cell.font = font_col
        cell.alignment = center
        cell.fill = fill_header
        cell.border = thin

    for r_i, row in enumerate(sheet.get("rows") or []):
        line = [
            row.get("code") or "",
            row.get("centre_name") or "",
            row.get("leader_name") or "",
            row.get("address") or "",
            row.get("contact") or "",
        ]
        for week in row.get("weeks") or []:
            line.extend(
                [
                    week.get("male_adults", 0),
                    week.get("female_adults", 0),
                    week.get("male_children", 0),
                    week.get("female_children", 0),
                    week.get("total", 0),
                    week.get("new_converts", 0),
                    week.get("first_timers", 0),
                    week.get("testimonies", 0),
                ]
            )
        ws.append(line)
        row_idx = ws.max_row
        row_fill = fill_alt if r_i % 2 else fill_white
        for col in range(1, last_col + 1):
            cell = ws.cell(row_idx, col)
            cell.font = font_cell
            cell.border = thin
            cell.fill = row_fill
            cell.alignment = center if col > meta_cols else left
        # T column highlight in each week block
        for i in range(len(labels)):
            t_col = meta_cols + 1 + i * week_width + 4
            ws.cell(row_idx, t_col).fill = fill_total_col

    if sheet.get("rows"):
        tot = ["TOTAL"] + [""] * (meta_cols - 1)
        for wt in sheet.get("week_totals") or []:
            tot.extend(
                [
                    wt.get("male_adults", 0),
                    wt.get("female_adults", 0),
                    wt.get("male_children", 0),
                    wt.get("female_children", 0),
                    wt.get("total", 0),
                    wt.get("new_converts", 0),
                    wt.get("first_timers", 0),
                    wt.get("testimonies", 0),
                ]
            )
        ws.append(tot)
        tot_idx = ws.max_row
        if meta_cols > 1:
            ws.merge_cells(
                start_row=tot_idx, start_column=1, end_row=tot_idx, end_column=meta_cols
            )
        for col in range(1, last_col + 1):
            cell = ws.cell(tot_idx, col)
            cell.font = font_totals
            cell.fill = fill_totals
            cell.border = thin
            cell.alignment = right if col == 1 else center

        grand = sheet.get("grand") or {}
        summaries = [
            ("GRAND TOTAL", grand.get("total", 0), font_grand),
            ("AVERAGE", sheet.get("average", 0), font_summary),
            ("TOTAL CELLS", sheet.get("total_cells", 0), font_summary),
            ("ACTIVE CELLS", sheet.get("active_cells", 0), font_summary),
        ]
        for label, value, font in summaries:
            ws.append([label, value])
            r = ws.max_row
            if last_col > 2:
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
            for col in range(1, last_col + 1):
                cell = ws.cell(r, col)
                cell.fill = fill_summary
                cell.border = thin
                cell.font = font
            ws.cell(r, 1).alignment = right
            ws.cell(r, 2).alignment = left

    ws.append([])
    ws.append(["ZONAL COORDINATOR"])
    ws.append(["Name:", "________________________________"])
    ws.append(["Signature:", "________________________________"])
    ws.append(["Date:", "________________________________"])

    # Identity columns stay readable; week stat columns share one width (MA = FA = …).
    ident_widths = {1: 10, 2: 20, 3: 18, 4: 20, 5: 14}
    for col in range(1, last_col + 1):
        letter = get_column_letter(col)
        if col <= meta_cols:
            ws.column_dimensions[letter].width = ident_widths.get(col, 14)
        else:
            ws.column_dimensions[letter].width = 6

    ws.freeze_panes = ws.cell(header_row_idx + 1, 3)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sheet_pdf(sheet: dict[str, Any]) -> bytes:
    """Build a landscape PDF summary of the attendance sheet."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleL",
        parent=styles["Heading1"],
        fontSize=14,
        alignment=0,
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "SubL",
        parent=styles["Normal"],
        fontSize=10,
        alignment=0,
        spaceAfter=2,
    )
    story = []
    org = sheet.get("organization_name") or sheet.get("assembly_name") or "Assembly"
    story.append(Paragraph(str(org), title))
    if sheet.get("branch_name"):
        story.append(Paragraph(str(sheet.get("branch_name")), sub))
    if sheet.get("show_zone") and sheet.get("zone_name"):
        story.append(Paragraph(str(sheet.get("zone_name")), sub))
    if sheet.get("show_subgroup") and sheet.get("subgroup_name"):
        story.append(Paragraph(str(sheet.get("subgroup_name")), sub))
    if sheet.get("report_date"):
        story.append(Paragraph(str(sheet.get("report_date")), sub))
    leader = sheet.get("leader_name") or sheet.get("coordinator_name") or "—"
    story.append(Paragraph(f"<b>LEADER NAME:</b> {leader}", sub))
    story.append(Spacer(1, 6))

    labels = sheet.get("week_labels") or []
    headers = ["CODE", "CENTRE", "LEADER", "ADDRESS"]
    for lab in labels:
        headers.append(f"{lab} T")
    data = [headers]
    for row in sheet.get("rows") or []:
        line = [
            str(row.get("code") or ""),
            str(row.get("centre_name") or "")[:28],
            str(row.get("leader_name") or "")[:22],
            str(row.get("address") or "")[:28],
        ]
        for week in row.get("weeks") or []:
            line.append(str(week.get("total", 0)))
        data.append(line)

    if sheet.get("rows"):
        tot = ["TOTAL", "", "", ""]
        for wt in sheet.get("week_totals") or []:
            tot.append(str(wt.get("total", 0)))
        data.append(tot)

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#64748B")),
                ("ALIGN", (4, 1), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (3, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 10))

    grand = sheet.get("grand") or {}
    summary = [
        ["GRAND TOTAL", str(grand.get("total", 0))],
        ["AVERAGE", str(sheet.get("average", 0))],
        ["TOTAL CELLS", str(sheet.get("total_cells", 0))],
        ["ACTIVE CELLS", str(sheet.get("active_cells", 0))],
    ]
    st = Table(summary, colWidths=[120, 80])
    st.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#64748B")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(st)
    story.append(Spacer(1, 16))
    story.append(Paragraph("<b>ZONAL COORDINATOR</b>", sub))
    story.append(Paragraph("Name: ________________________________", sub))
    story.append(Paragraph("Signature: ____________________________", sub))
    story.append(Paragraph("Date: ________________________________", sub))

    doc.build(story)
    return buf.getvalue()
