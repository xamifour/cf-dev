# cf-dev/cf_src/appsinn/cf_operations/attendance_import.py

"""
Import zonal attendance sheets matching Attendence.xlsx layout.

Creates AttendanceRecord rows (source of truth) with optional week labels
and AttendanceSeat headcounts. Only weeks that have counts on the sheet
become records — weeks 1–5 are never pre-populated as empty rows.
Reports are generated on the fly from those records.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import BinaryIO

from django.db import transaction

from .models import AttendanceRecord, AttendanceSeat, Event

# Week blocks start at 1-based Excel columns: F=6, N=14, V=22, AD=30, AL=38
_WEEK_START_COLS = (6, 14, 22, 30, 38)
_WEEK_FIELD_OFFSETS = {
    0: "male_adults",  # MA
    1: "female_adults",  # FA
    2: "male_children",  # MC
    3: "female_children",  # FC
    4: "total",  # T (recomputed on save)
    5: "new_converts",  # N/C
    6: "first_timers",  # F/T
    7: "testimonies",  # TS
}
_STOP_ROW_MARKERS = {
    "TOTAL",
    "GRAND TOTAL",
    "AVERAGE",
    "TOTAL CELLS",
    "ACTIVE CELLS",
    "LEGEND",
    "LEGEND:",
}


@dataclass
class ImportResult:
    event: Event | None = None
    records_created: int = 0
    seats_created: int = 0
    assembly_name: str = ""
    zone_name: str = ""
    report_title: str = ""
    coordinator_name: str = ""
    week_labels: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    record_ids: list = field(default_factory=list)


def _cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def _as_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _as_int(value) -> int:
    if value is None or value == "":
        return 0
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def _week_has_counts(week_data: dict) -> bool:
    return any(int(week_data.get(k) or 0) for k in week_data)


def _read_bytes(file_obj: BinaryIO) -> bytes:
    """
    Materialise upload bytes so openpyxl works with Django TemporaryUploadedFile
    and other non-seekable / already-read streams.
    """
    if hasattr(file_obj, "seek"):
        try:
            file_obj.seek(0)
        except Exception:  # noqa: BLE001
            pass
    if hasattr(file_obj, "read"):
        data = file_obj.read()
    else:
        data = file_obj
    if isinstance(data, memoryview):
        data = data.tobytes()
    if isinstance(data, bytearray):
        data = bytes(data)
    if not isinstance(data, (bytes, bytearray)):
        raise ValueError("Expected a binary Excel upload (.xlsx).")
    if not data:
        raise ValueError("The uploaded file is empty.")
    # Minimal zip/xlsx signature check (PK..)
    if not data.startswith(b"PK"):
        raise ValueError(
            "File does not look like a .xlsx workbook. "
            "Save as Excel Workbook (.xlsx) and try again."
        )
    return bytes(data)


def parse_attendance_workbook(file_obj: BinaryIO) -> dict:
    """Parse an Attendence.xlsx-style workbook into a plain dict (no DB)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "openpyxl is required to import attendance sheets. "
            "Install it with: pip install openpyxl"
        ) from exc

    data = _read_bytes(file_obj)
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active sheet.")

        assembly_name = _as_str(_cell(ws, 2, 1)) or "Assembly"
        report_title = _as_str(_cell(ws, 3, 1)) or "Zonal Report"
        zone_name = _as_str(_cell(ws, 4, 1)) or "Zone"

        coord_raw = _as_str(_cell(ws, 5, 2))
        coordinator_name = coord_raw
        for prefix in (
            "ZONAL CORDINATOR NAME:",
            "ZONAL COORDINATOR NAME:",
            "ZONAL COORDINATOR NAME",
            "ZONAL CORDINATOR NAME",
        ):
            if coordinator_name.upper().startswith(prefix):
                coordinator_name = coordinator_name[len(prefix) :].strip(" :")
                break

        week_labels = []
        for col in _WEEK_START_COLS:
            label = _as_str(_cell(ws, 5, col))
            week_labels.append(label or f"WEEK {len(week_labels) + 1}")

        centres = []
        max_row = ws.max_row or 7
        # read_only max_row can be None / huge; cap scan
        if max_row is None or max_row > 5000:
            max_row = 500

        for row in range(7, max_row + 1):
            centre_name = _as_str(_cell(ws, row, 2))
            marker_f = _as_str(_cell(ws, row, 6)).upper()
            marker_a = _as_str(_cell(ws, row, 1)).upper()
            marker_b = centre_name.upper()
            if (
                marker_f in _STOP_ROW_MARKERS
                or marker_a in _STOP_ROW_MARKERS
                or marker_b in _STOP_ROW_MARKERS
            ):
                break
            if not centre_name and _cell(ws, row, 1) is None:
                # Skip fully empty rows but keep scanning a bit.
                if not any(_cell(ws, row, c) for c in range(1, 14)):
                    continue
            if not centre_name:
                continue

            # Column A is optional code (legacy S/N); keep as text.
            code_raw = _cell(ws, row, 1)
            code = _as_str(code_raw)
            if not code and code_raw is not None:
                code = str(_as_int(code_raw) or "")
            if not code:
                code = str(len(centres) + 1)
            weeks = []
            for start_col in _WEEK_START_COLS:
                stat = {}
                for offset, field_name in _WEEK_FIELD_OFFSETS.items():
                    if field_name == "total":
                        continue
                    stat[field_name] = _as_int(_cell(ws, row, start_col + offset))
                weeks.append(stat)
            centres.append(
                {
                    "code": code,
                    "centre_name": centre_name,
                    "leader": _as_str(_cell(ws, row, 3)),
                    "address": _as_str(_cell(ws, row, 4)),
                    "weeks": weeks,
                }
            )
    finally:
        wb.close()

    return {
        "assembly_name": assembly_name,
        "report_title": report_title,
        "zone_name": zone_name,
        "coordinator_name": coordinator_name,
        "week_labels": week_labels,
        "centres": centres,
    }


def _resolve_zone(branch, zone_name: str):
    """Match a Zone under the branch by name when possible."""
    if not zone_name or not branch:
        return None
    try:
        from cf_people.models import Zone
    except Exception:  # noqa: BLE001
        return None
    name = zone_name.strip()
    qs = Zone.objects.filter(branch=branch)
    zone = qs.filter(name__iexact=name).first()
    if zone:
        return zone
    # Tolerate "ZONE 13" vs "Zone 13 - Kasoa"
    return qs.filter(name__icontains=name).order_by("name").first()


@transaction.atomic
def import_attendance_from_excel(
    *,
    branch,
    file_obj: BinaryIO,
    event: Event | None = None,
    created_by=None,
) -> ImportResult:
    """
    Import sheet rows as AttendanceRecord + optional AttendanceSeat.

    For each centre, only weeks that have non-zero counts create a record
    (with ``week`` set and a seat). Empty week columns are skipped — no
    week1–week5 pre-population.
    """
    if branch is None:
        raise ValueError("Branch is required.")

    parsed = parse_attendance_workbook(file_obj)
    result = ImportResult(
        assembly_name=parsed["assembly_name"],
        zone_name=parsed["zone_name"],
        report_title=parsed["report_title"],
        coordinator_name=parsed["coordinator_name"],
        week_labels=parsed["week_labels"],
        warnings=[],
    )

    if not parsed["centres"]:
        raise ValueError(
            "No centre / cell rows found in the sheet. "
            "Expected names in column B starting at row 7."
        )

    if event is not None and event.branch_id != branch.pk:
        raise ValueError("Selected event belongs to a different branch.")

    if event is None:
        title = (parsed["report_title"] or "Imported attendance").strip()[:255]
        desc_bits = [
            f"Imported attendance sheet for {parsed.get('assembly_name') or branch}.",
        ]
        if parsed.get("zone_name"):
            desc_bits.append(f"Zone: {parsed['zone_name']}.")
        if parsed.get("coordinator_name"):
            desc_bits.append(f"Coordinator: {parsed['coordinator_name']}.")
        if parsed.get("week_labels"):
            desc_bits.append("Weeks: " + "; ".join(parsed["week_labels"][:5]) + ".")
        event = Event.objects.create(
            branch=branch,
            title=title,
            event_type="SERVICE",
            visibility=Event.VISIBILITY_BRANCH,
            description="\n".join(desc_bits),
            created_by=created_by,
            modified_by=created_by,
        )
    result.event = event

    zone = _resolve_zone(branch, parsed.get("zone_name") or "")
    if parsed.get("zone_name") and zone is None:
        result.warnings.append(
            f"No matching Zone named “{parsed['zone_name']}” under this branch; "
            "records were imported without a zone link."
        )

    for centre in parsed["centres"]:
        filled_weeks = [
            (week_num, week_data)
            for week_num, week_data in enumerate(centre["weeks"], start=1)
            if week_num <= 5 and _week_has_counts(week_data)
        ]

        base_kwargs = dict(
            event=event,
            branch=branch,
            zone=zone,
            code=centre.get("code") or "",
            centre_name=centre["centre_name"],
            leader=centre.get("leader") or "",
            address=centre.get("address") or "",
            created_by=created_by,
            modified_by=created_by,
        )

        if not filled_weeks:
            # Centre listed with no counts: still create identity row (no seat).
            record = AttendanceRecord.objects.create(**base_kwargs)
            result.records_created += 1
            result.record_ids.append(record.pk)
            continue

        for week_num, week_data in filled_weeks:
            record = AttendanceRecord.objects.create(week=week_num, **base_kwargs)
            result.records_created += 1
            result.record_ids.append(record.pk)

            AttendanceSeat.objects.create(
                record=record,
                male_adults=week_data.get("male_adults", 0),
                female_adults=week_data.get("female_adults", 0),
                male_children=week_data.get("male_children", 0),
                female_children=week_data.get("female_children", 0),
                new_converts=week_data.get("new_converts", 0),
                first_timers=week_data.get("first_timers", 0),
                testimonies=week_data.get("testimonies", 0),
            )
            result.seats_created += 1

    return result
