# cf-dev/cf_src/appsinn/cf_operations/attendance_export.py

"""Export attendance sheet dict to Excel (.xlsx) or PDF."""

from __future__ import annotations

from io import BytesIO
from typing import Any


def export_sheet_xlsx(sheet: dict[str, Any]) -> bytes:
    """Build an Excel workbook matching the preview grid."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    labels = sheet.get("week_labels") or [
        "WEEK 1",
        "WEEK 2",
        "WEEK 3",
        "WEEK 4",
        "WEEK 5",
    ]
    # Header rows: org / branch / zone (if) / subgroup (if) / date
    ws.append([sheet.get("organization_name") or sheet.get("assembly_name") or ""])
    if sheet.get("branch_name"):
        ws.append([sheet.get("branch_name")])
    if sheet.get("show_zone") and sheet.get("zone_name"):
        ws.append([sheet.get("zone_name")])
    if sheet.get("show_subgroup") and sheet.get("subgroup_name"):
        ws.append([sheet.get("subgroup_name")])
    if sheet.get("report_date"):
        ws.append([sheet.get("report_date")])
    leader = sheet.get("leader_name") or sheet.get("coordinator_name") or ""
    ws.append([f"LEADER NAME: {leader}"])

    col_headers = ["CODE", "CENTRE NAME", "LEADER'S NAME", "ADDRESS", "CONTACT"]
    for lab in labels:
        col_headers.extend(["MA", "FA", "MC", "FC", "T", "N/C", "F/T", "TS"])
    # Week labels row
    week_row = ["", "", "", "", ""]
    for lab in labels:
        week_row.append(lab)
        week_row.extend([""] * 7)
    ws.append(week_row)
    ws.append(col_headers)

    header_font = Font(bold=True)
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.font = header_font
    left = Alignment(horizontal="left", vertical="center")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1):
        for cell in row:
            cell.alignment = left

    for row in sheet.get("rows") or []:
        line = [
            row.get("code") or "",
            row.get("centre_name") or "",
            row.get("leader_name") or "",
            row.get("address") or "",
            row.get("contact") or "",
        ]
        for week in row.get("weeks") or []:
            line.extend(
                [
                    week.get("male_adults", 0),
                    week.get("female_adults", 0),
                    week.get("male_children", 0),
                    week.get("female_children", 0),
                    week.get("total", 0),
                    week.get("new_converts", 0),
                    week.get("first_timers", 0),
                    week.get("testimonies", 0),
                ]
            )
        ws.append(line)

    # Totals
    if sheet.get("rows"):
        tot = ["TOTAL", "", "", "", ""]
        for wt in sheet.get("week_totals") or []:
            tot.extend(
                [
                    wt.get("male_adults", 0),
                    wt.get("female_adults", 0),
                    wt.get("male_children", 0),
                    wt.get("female_children", 0),
                    wt.get("total", 0),
                    wt.get("new_converts", 0),
                    wt.get("first_timers", 0),
                    wt.get("testimonies", 0),
                ]
            )
        ws.append(tot)
        grand = sheet.get("grand") or {}
        summary_start = ws.max_row + 1
        ws.append(["GRAND TOTAL", grand.get("total", 0)])
        ws.append(["AVERAGE", sheet.get("average", 0)])
        ws.append(["TOTAL CELLS", sheet.get("total_cells", 0)])
        ws.append(["ACTIVE CELLS", sheet.get("active_cells", 0)])
        for r in range(summary_start, ws.max_row + 1):
            ws.cell(r, 1).alignment = Alignment(horizontal="right")
            ws.cell(r, 2).alignment = Alignment(horizontal="left")

    ws.append([])
    ws.append(["ZONAL COORDINATOR"])
    ws.append(["Name:", "________________________________"])
    ws.append(["Signature:", "________________________________"])
    ws.append(["Date:", "________________________________"])

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[letter].width = min(28, max(8, max_len + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sheet_pdf(sheet: dict[str, Any]) -> bytes:
    """Build a landscape PDF summary of the attendance sheet."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "TitleL",
        parent=styles["Heading1"],
        fontSize=14,
        alignment=0,
        spaceAfter=4,
    )
    sub = ParagraphStyle(
        "SubL",
        parent=styles["Normal"],
        fontSize=10,
        alignment=0,
        spaceAfter=2,
    )
    story = []
    org = sheet.get("organization_name") or sheet.get("assembly_name") or "Assembly"
    story.append(Paragraph(str(org), title))
    if sheet.get("branch_name"):
        story.append(Paragraph(str(sheet.get("branch_name")), sub))
    if sheet.get("show_zone") and sheet.get("zone_name"):
        story.append(Paragraph(str(sheet.get("zone_name")), sub))
    if sheet.get("show_subgroup") and sheet.get("subgroup_name"):
        story.append(Paragraph(str(sheet.get("subgroup_name")), sub))
    if sheet.get("report_date"):
        story.append(Paragraph(str(sheet.get("report_date")), sub))
    leader = sheet.get("leader_name") or sheet.get("coordinator_name") or "—"
    story.append(Paragraph(f"<b>LEADER NAME:</b> {leader}", sub))
    story.append(Spacer(1, 6))

    # Compact table: code, centre, leader, address + week totals only (T column)
    headers = ["CODE", "CENTRE", "LEADER", "ADDRESS"]
    labels = sheet.get("week_labels") or ["W1", "W2", "W3", "W4", "W5"]
    for i, lab in enumerate(labels[:5], start=1):
        headers.append(f"W{i} T")
    data = [headers]
    for row in sheet.get("rows") or []:
        line = [
            str(row.get("code") or ""),
            str(row.get("centre_name") or "")[:28],
            str(row.get("leader_name") or "")[:22],
            str(row.get("address") or "")[:28],
        ]
        for week in (row.get("weeks") or [])[:5]:
            line.append(str(week.get("total", 0)))
        data.append(line)

    if sheet.get("rows"):
        tot = ["TOTAL", "", "", ""]
        for wt in (sheet.get("week_totals") or [])[:5]:
            tot.append(str(wt.get("total", 0)))
        data.append(tot)

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (4, 1), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (3, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 10))

    grand = sheet.get("grand") or {}
    summary = [
        ["GRAND TOTAL", str(grand.get("total", 0))],
        ["AVERAGE", str(sheet.get("average", 0))],
        ["TOTAL CELLS", str(sheet.get("total_cells", 0))],
        ["ACTIVE CELLS", str(sheet.get("active_cells", 0))],
    ]
    st = Table(summary, colWidths=[120, 80])
    st.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(st)
    story.append(Spacer(1, 16))
    story.append(Paragraph("<b>ZONAL COORDINATOR</b>", sub))
    story.append(Paragraph("Name: ________________________________", sub))
    story.append(Paragraph("Signature: ____________________________", sub))
    story.append(Paragraph("Date: ________________________________", sub))

    doc.build(story)
    return buf.getvalue()
